# utils/exportar_dict_a_excel.py - Versión sin pandas
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

def exportar_dict_a_excel(data, filename=None):
    """
    Exporta diccionario a Excel usando openpyxl en lugar de pandas
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Si data es una lista de diccionarios
    if isinstance(data, list) and len(data) > 0:
        # Headers
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Data rows
        for row_idx, item in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=item.get(header, ""))
    
    # Si data es un diccionario simple
    elif isinstance(data, dict):
        ws.cell(row=1, column=1, value="Campo").font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=2, value="Valor").font = header_font
        ws.cell(row=1, column=2).fill = header_fill
        
        for row_idx, (key, value) in enumerate(data.items(), 2):
            ws.cell(row=row_idx, column=1, value=str(key))
            ws.cell(row=row_idx, column=2, value=str(value))
    
    # Ajustar columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar en BytesIO para envío
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer